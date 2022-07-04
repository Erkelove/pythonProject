import unittest

import openpyxl

from UTEST.my_unit import unite

from ddt import ddt, data, unpack


def read_excel():
    book = openpyxl.load_workbook('data.xlsx')
    sh1 = book['login']
    #print(sh1.max_row, sh1.max_column)

    alllist = []
    for row in range(2, sh1.max_row + 1):
        templist = []
        for col in range(1, sh1.max_column + 1):
            # print(row, col)
            templist.append(sh1.cell(row, col).value)
        #print(templist)
        alllist.append(templist)
    #print(alllist)
    return alllist

@ddt
class test_print01(unite):

    age = 18
    @unittest.skip
    def test01(self):
        print("测试知了")

    @unittest.skipIf(age > 16, "微微大于16周岁了，可以跳过")
    def test02(self):
        print("测试微微")

    @unittest.skipUnless(age < 17, "百里没小于17周岁，跳过")
    def test03(self):
        print("测试百里")

    @data(*read_excel())
    #@unpack
    def test04(self, args):
        print(args)
        print('-----------')

